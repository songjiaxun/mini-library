# -*- coding: utf-8 -*-
from openpyxl import load_workbook

reader = "读者信息.xlsx"
header1 = "==============================================================================="
try:
    wr2 = load_workbook(reader)
    ws2 = wr2.worksheets[0]
except Exception:
    input("请确认“图书馆信息.xlsx”和“读者信息.xlsx”文件保持关闭状态，并与该软件置于同一目录下！")

def reader_id_generater():
    ###自动生成借书号###
    print (header1)
    grade = {
            "一": "1",
            "二": "2",
            "三": "3",
            "四": "4",
            "五": "5",
            "六": "6",
            "七": "7",
            "八": "8",
            "九": "9"
            }    
    class_num = {
            "年级": "0",
            "甲": "1",
            "乙": "2",
            "丙": "3",
            "丁": "4",
            "一": "1",
            "二": "2",
            "三": "3",
            "四": "4",
            "五": "5",
            "六": "6",
            "七": "7",
            "八": "8",
            "九": "9", 
            "十": "10"            
            }
    nrows = ws2.max_row
    classid_dic = {}
    for row in range(2,nrows+1):
        if ws2.cell(row=row,column=4).value != "教师":
            reader_id = ""
            classid = ws2.cell(row=row,column=4).value
            if classid in classid_dic.keys():
                classid_dic[classid] += 1
            else:
                classid_dic[classid] = 1
            count = classid_dic[classid]
            for g in grade:
                if classid[0].find(g) != -1:
                    reader_id += grade[g]
                    break
            for n in class_num:
                if classid[1:].find(n) != -1:
                    reader_id += class_num[n]
                    break
            if count < 10:
                count_str = "0" + str(count)
            else:
                count_str = str(count)
            reader_id += count_str
            ws2.cell(row=row,column=1).value = int(reader_id)
        elif ws2.cell(row=row,column=4).value == "教师":
            classid = ws2.cell(row=row,column=4).value
            if classid in classid_dic.keys():
                classid_dic[classid] += 1
            else:
                classid_dic[classid] = 1
            count = classid_dic[classid]
            ws2.cell(row=row,column=1).value = count
    wr2.save(reader)
