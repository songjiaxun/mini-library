# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import json

library = "图书馆信息.xlsx"
reader = "读者信息.xlsx"
data_book = {}
data_reader = {}
header1 = "==============================================================================="
header2 = "-------------------------------------------------------------------------------"
###从两个excel文档和一个json文档中载入初始数据###
try:
    wb_lib = load_workbook(library)
    wb_reader = load_workbook(reader)
    ws_book = wb_lib.worksheets[0]#书籍信息
    ws_reader = wb_reader.worksheets[0]#读者信息
    ws_lostbook = wb_lib.worksheets[1]#丢失书籍信息
    ws_log = wb_reader.worksheets[1]#借阅记录
except Exception:
    input("请确认“图书馆信息.xlsx”和“读者信息.xlsx”文件保持关闭状态，并与该软件置于同一目录下！")

try:
    with open('meta_data.json','r') as f:
        json_data = json.load(f)
        supposed_return_days_students = json_data['student_days']
        supposed_return_days_teachers = json_data['teacher_days']
except Exception:
    input('请确认‘meta_data.json’文件保持关闭状态，并与该软件置于同一目录下！')

def bookinfo_read():
    ###从excel文件读取图书信息###
    print('将read_bookinfo方法修改成了bookinfo_read方法')

def initiallize():
    ###初始化###

    if json_data['status'] == '0':
        print('【软件初始化】，请按提示输入相应内容：')
        json_data['status'] = '1'
        json_data['institution'] = input('【学校/机构名称】：')
        json_data['password'] = input('【登陆密码】：')
        json_data['administrator'] = input('【管理员密码】：')
        with open('meta_data.json', 'w') as meta_w:
            json.dump(json_data,meta_w)
    return (json_data['institution'], json_data['password'], json_data['administrator'])



institution, pw, pw_admin = initiallize()
bookinfo_read()
