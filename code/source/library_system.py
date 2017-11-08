# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import datetime
import json
import copy
from id_generater import *
from library_data_entry_excel import *

library = "图书馆信息.xlsx"
reader = "读者信息.xlsx"
data_book = {}
data_reader = {}
header1 = "==============================================================================="
header2 = "-------------------------------------------------------------------------------"
###从两个excel文档和一个json文档中载入初始数据###
try:
    wr = load_workbook(library)
    wr2 = load_workbook(reader)
    ws = wr.worksheets[0]
    ws2 = wr2.worksheets[0]
    ws3 = wr.worksheets[1]
    ws4 = wr2.worksheets[1]
except Exception:
    input("请确认“图书馆信息.xlsx”和“读者信息.xlsx”文件保持关闭状态，并与该软件置于同一目录下！")

try:
    with open('meta_data.json', "r") as f:
        json_data = json.load(f)
        supposed_return_days_students = json_data["student_days"]
        supposed_return_days_teachers = json_data["teacher_days"]
except Exception:
    input("请确认“meta_data.json”文件保持关闭状态，并与该软件置于同一目录下！")

def read_bookinfo():
    ###从excel文件读取图书信息###
    nrows = ws.max_row
    for book in range(2, nrows+1):
        if ws.cell(row=book,column=13).value == None:
            borrowed_times = 0
        else:
            borrowed_times = ws.cell(row=book,column=13).value
        if not str(ws.cell(row=book,column=1).value).isdigit():
            isbn = "0"
        else:
            isbn = str(int(ws.cell(row=book,column=1).value))
        if not str(ws.cell(row=book,column=9).value).isdigit():
            number = 0
        else:
            number = int(ws.cell(row=book,column=9).value)
        bookinfo = {
                "row": book,
                "ISBN": isbn,
                "书籍名称": ws.cell(row=book,column=2).value,
                "作者": ws.cell(row=book,column=3).value,
                "出版社": ws.cell(row=book,column=4).value,
                "出版日期": ws.cell(row=book,column=5).value,
                "页数": ws.cell(row=book,column=6).value,
                "价格": ws.cell(row=book,column=7).value,
                "主题": ws.cell(row=book,column=8).value,
                "馆藏本数": number,
                "索书号": ws.cell(row=book,column=10).value,
                "书籍位置": ws.cell(row=book,column=15).value,
                "借阅次数": borrowed_times,
                "借阅记录": ws.cell(row=book,column=14).value,
                "内容简介": ws.cell(row=book,column=11).value,
                "信息来源": ws.cell(row=book,column=12).value
                }
        data_book[isbn] = bookinfo

def read_readerinfo():
    ###从excel文件读取读者信息###
    nrows = ws2.max_row    
    for r in range(2, nrows+1):
        if ws2.cell(row=r,column=1).value == None:
            continue
        if ws2.cell(row=r,column=11).value == None:
            count = 0
        else:
            count = ws2.cell(row=r,column=11).value
        if ws2.cell(row=r,column=12).value == None:
            borrowed_times = 0
        else:
            borrowed_times = ws2.cell(row=r,column=12).value
        if not str(ws2.cell(row=r,column=1).value).isdigit():
            reader_id = 0
        else:
            reader_id = str(int(ws2.cell(row=r,column=1).value))
        readerinfo = {
                "row": r,
                "借书号": reader_id,
                "姓名": ws2.cell(row=r,column=2).value,
                "性别": ws2.cell(row=r,column=3).value,
                "单位": ws2.cell(row=r,column=4).value,
                "所借书目": ws2.cell(row=r,column=5).value,
                "借书日期": ws2.cell(row=r,column=6).value,
                "应还日期": ws2.cell(row=r,column=7).value,
                "还书日期": ws2.cell(row=r,column=8).value,
                "借书记录": ws2.cell(row=r,column=9).value,
                "借书权限": ws2.cell(row=r,column=10).value,
                "过期次数": count,
                "借阅次数": borrowed_times
                }                
        data_reader[reader_id] = readerinfo

def sort_borrow_log(d):
    ###生成借书记录###
    log = {}
    for reader_id in d:
        request = d[reader_id]
        if request["借阅次数"] == 0:
            pass
        else:
            borrow_log = request["借书记录"].split(",")
            for book in borrow_log:
                if book.find("(") != -1:
                    log[datetime.datetime.strptime(book[book.find("(")+1:-1], "%Y-%m-%d %H:%M:%S")] = [request["单位"], request["姓名"], request["借书号"], "借书", book[:book.find("(")], data_book[book[:book.find("(")]]["书籍名称"], data_book[book[:book.find("(")]]["书籍位置"]]
                elif book.find("{") != -1:
                    log[datetime.datetime.strptime(book[book.find("{")+1:-1], "%Y-%m-%d %H:%M:%S")] = [request["单位"], request["姓名"], request["借书号"], "丢书", book[:book.find("{")], data_book[book[:book.find("{")]]["书籍名称"], data_book[book[:book.find("{")]]["书籍位置"]]
                elif book.find("[") != -1:
                    log[datetime.datetime.strptime(book[book.find("[")+1:-1], "%Y-%m-%d %H:%M:%S")] = [request["单位"], request["姓名"], request["借书号"], "还书", book[:book.find("[")], data_book[book[:book.find("[")]]["书籍名称"], data_book[book[:book.find("[")]]["书籍位置"]]
    items=log.items()
    backitems=[[v[0],v[1]] for v in items]
    backitems.sort(reverse=True)
    return log, [backitems[i][0] for i in range(0,len(backitems))]

def summary():
    ###统计馆藏本书、注册读者数等信息###
    global summary1, summary2, summary3
    summary1 = 0
    summary2 = 0
    summary3 = 0
    for book in data_book:
        if data_book[book]["馆藏本数"] == None:
            pass
        else:
            summary1 += data_book[book]["馆藏本数"]
            summary2 += 1            
    for reader in data_reader:
        summary3 += 1

def info_summary (data_book):
    ###生成管理信息###
    def sort_value(d):
        ###排序模块###
        items=d.items()
        for v in items:
            if type(v[0]) == int:
                print (v[0],type(v[0]))
        backitems=[[v[1]["借阅次数"],v[0]] for v in items]
        backitems.sort(reverse=True)
        return [backitems[i][1] for i in range(0,len(backitems))]
    print (header1)
    summary()
    print ("图书馆现存图书【"+str(summary2)+"】种，共计图书【"+str(summary1)+"】册，注册读者【"+str(summary3)+"】人。")
    print (header2)
    print ("最受欢迎的10本图书：")
    ranking = 1
    for book in sort_value(data_book)[:10]:
        print ("【"+str(ranking)+"】共借阅"+str(data_book[book]["借阅次数"])+"次，书籍位置："+str(data_book[book]["书籍位置"])+"，书籍名称："+str(data_book[book]["书籍名称"]))
        ranking += 1
    print (header2)
    print ("最勤奋的10位读者：")
    ranking = 1
    for r in sort_value(data_reader)[:10]:
        print ("【"+str(ranking)+"】共借阅"+str(data_reader[r]["借阅次数"])+"次，单位："+str(data_reader[r]["单位"])+"，姓名："+str(data_reader[r]["姓名"]))
        ranking += 1
    print (header2)
    print ("过期未还书的读者：")
    for reader in data_reader:
        if data_reader[reader]["应还日期"] == None:
            pass
        elif data_reader[reader]["应还日期"]<datetime.datetime.now():
            print (data_reader[reader]["借书号"], data_reader[reader]["单位"], data_reader[reader]["姓名"], data_reader[reader]["所借书目"], data_book[data_reader[reader]["所借书目"]]["书籍名称"], "应还日期:", data_reader[reader]["应还日期"].strftime('%Y-%m-%d'))
    print (header2)
    print ("今日借阅记录（查看详细借阅记录请打开“读者信息.xlsx”文件查询）：")
    borrow_log, sorted_log = sort_borrow_log(data_reader)
    r = 2
    for log in sorted_log:
        if log.date() == datetime.datetime.now().date():
            print (log, borrow_log[log][0], borrow_log[log][1], borrow_log[log][2], borrow_log[log][3], borrow_log[log][4], borrow_log[log][5], borrow_log[log][6])
        ws4.cell(row=r,column=1).value = log
        ws4.cell(row=r,column=2).value = borrow_log[log][0]
        ws4.cell(row=r,column=3).value = borrow_log[log][1]
        ws4.cell(row=r,column=4).value = borrow_log[log][2]
        ws4.cell(row=r,column=5).value = borrow_log[log][3]
        ws4.cell(row=r,column=6).value = borrow_log[log][4]
        ws4.cell(row=r,column=7).value = borrow_log[log][5]
        ws4.cell(row=r,column=8).value = borrow_log[log][6]
        r += 1    
    wr2.save("读者信息.xlsx")
    print (header1)
    input ("点击回车键确认退出。")

def input_request (instruction):
    ###信息输入###
    content = input(instruction)
    while content == "":
        content = input(instruction)
    return content

def borrow_book():
    ###借书模块###
    print (header1)
    readerid = input_request("请输入读者【借书号】，退出借书请按【0】\n读者借书号:")
    while readerid != "0":
        try:
            request_reader = data_reader[readerid]
            print_request_reader(request_reader, False)
            if request_reader["借书权限"] != "开通":
                print (header2+"\n【该读者暂无借书权限！】")
                input(header2+"\n点击回车键确认退出借书。")
                return
            if request_reader["所借书目"] != None:
                print (header2+"\n【请将已借书目归还后再借新书！】")
                input(header2+"\n点击回车键确认退出借书。")
                return
            print (header2)
            isbn = input_request("请输入欲借书目【ISBN】，退出借书请按【0】\nISBN:")
            if isbn == "0":
                return
            while isbn != "0":
                try:
                    request_book = data_book[isbn]
                    print_request_book(request_book, False)
                    print (header2)
                    confirm = input_request("是否确认借书？确认请按【1】，取消请按【0】")
                    while confirm != "1" and confirm != "0":
                        confirm = input_request("是否确认借书？确认请按【1】，取消请按【0】")
                    if confirm == "1":
                        #修改读者信息
                        ws2.cell(row=request_reader["row"],column=5).value = isbn
                        #借书时间计算
                        now = datetime.datetime.now()
                        if request_reader["单位"] == "教师":
                            return_time = now + datetime.timedelta(days=int(supposed_return_days_teachers))
                        else:
                            return_time = now + datetime.timedelta(days=int(supposed_return_days_students))
                        now_str = now.strftime('%Y-%m-%d %H:%M:%S')
                        ws2.cell(row=request_reader["row"],column=6).value = now
                        ws2.cell(row=request_reader["row"],column=7).value = return_time
                        if ws2.cell(row=request_reader["row"],column=9).value == None:
                            ws2.cell(row=request_reader["row"],column=9).value = isbn + "(" + now_str + ")"
                        else:
                            ws2.cell(row=request_reader["row"],column=9).value = request_reader["借书记录"] + "," + isbn + "(" + now_str + ")"
                        if ws2.cell(row=request_reader["row"],column=12).value == None:
                            ws2.cell(row=request_reader["row"],column=12).value = 1
                        else:
                            ws2.cell(row=request_reader["row"],column=12).value += 1
                        wr2.save(reader)
                        #修改图书信息
                        if ws.cell(row=request_book["row"],column=13).value == None:
                            ws.cell(row=request_book["row"],column=13).value = 1
                        else:
                            ws.cell(row=request_book["row"],column=13).value += 1
                        if ws.cell(row=request_book["row"],column=14).value == None:
                            ws.cell(row=request_book["row"],column=14).value = "(" + now_str + ")" + request_reader["借书号"] + request_reader["单位"] + request_reader["姓名"] + "[借书]"
                        else:
                            ws.cell(row=request_book["row"],column=14).value += "\n          " + "(" + now_str + ")" + request_reader["借书号"] + request_reader["单位"] + request_reader["姓名"] + "[借书]"
                        wr.save(library)
                        read_bookinfo()
                        read_readerinfo()                        
                        print ("\n【借书成功！】")
                        input(header2+"\n点击回车键确认退出借书。")
                        return
                    else:
                        return
                except Exception:
                    print ("\n【ISBN不存在，请重新输入！】")
                    print (header2)
                    isbn = input_request("请输入欲借书目【ISBN】，退出借书请按【0】\nISBN:")
                    if isbn == "0":
                        return
        except Exception:
            print ("\n【读者借书号不存在，请重新输入！】")
            print (header2)
            readerid = input_request("请输入读者【借书号】，退出借书请按【0】\n读者借书号:")

def return_book():
    ###还书模块###
    print (header1)
    readerid = input_request("请输入读者【借书号】，退出还书请按【0】\n读者借书号:")
    while readerid != "0":
        try:
            request_reader = data_reader[readerid]
            print_request_reader(request_reader, False)
            if request_reader["所借书目"] == None:
                print (header2+"\n【该读者暂无借阅书目！】")
                input(header2+"\n点击回车键确认退出。")
                return
            print (header2)
            isbn = input_request("请输入欲还书目【ISBN】，书目丢失请输入【9999】，退出还书请按【0】\nISBN/9999:")
            while isbn != request_reader["所借书目"]:
                if isbn == "0":
                    return
                if isbn == "9999":
                    borrowed_book = request_reader["所借书目"]
                    request_book = data_book[borrowed_book]
                    now = datetime.datetime.now()
                    now_str = now.strftime('%Y-%m-%d %H:%M:%S')
                    #修改读者信息
                    ws2.cell(row=request_reader["row"],column=9).value = request_reader["借书记录"] + "," + borrowed_book  + "{" + now_str + "}"
                    ws2.cell(row=request_reader["row"],column=10).value = "丢失"
                    ws2.cell(row=request_reader["row"],column=5).value = None
                    ws2.cell(row=request_reader["row"],column=6).value = None
                    ws2.cell(row=request_reader["row"],column=7).value = None
                    ws2.cell(row=request_reader["row"],column=8).value = now
                    wr2.save(reader)
                    #添加丢书读者信息                    
                    nrows = ws3.max_row                   
                    ws3.cell(row=nrows+1,column=1).value = request_reader["单位"]
                    ws3.cell(row=nrows+1,column=2).value = request_reader["姓名"]
                    ws3.cell(row=nrows+1,column=3).value = request_reader["性别"]
                    ws3.cell(row=nrows+1,column=4).value = request_book["ISBN"]
                    ws3.cell(row=nrows+1,column=5).value = request_book["书籍名称"]
                    ws3.cell(row=nrows+1,column=6).value = request_book["价格"]
                    ws3.cell(row=nrows+1,column=7).value = now
                    #修改图书信息
                    ws.cell(row=request_book["row"],column=9).value -= 1
                    ws.cell(row=request_book["row"],column=14).value += "\n          " + "(" + now_str + ")" + request_reader["借书号"]  + request_reader["单位"] + request_reader["姓名"] + "[丢失]"
                    wr.save(library)                    
                    read_bookinfo()
                    read_readerinfo()
                    print (header2+"\n【读者借书权限已经关闭！】\n请赔偿所丢书目" + data_book[request_reader["所借书目"]]["书籍名称"] + "双倍价格！书籍价格为" + data_book[request_reader["所借书目"]]["价格"])
                    input(header2+"\n点击回车键确认退出。")
                    return
                isbn = input_request(header2+"\n【欲还书目与所借书目不符】，请输入欲还书目【ISBN】，书目丢失请输入【9999】，退出还书请按【0】\nISBN/9999:")          
            confirm = input_request("是否确认还书？确认请按【1】，取消请按【0】")
            while confirm != "1" and confirm != "0":
                confirm = input_request("是否确认还书？确认请按【1】，取消请按【0】")
            if confirm == "1":                
                #还书时间计算
                borrowed_book = request_reader["所借书目"]
                request_book = data_book[borrowed_book]
                now = datetime.datetime.now()
                now_str = now.strftime('%Y-%m-%d %H:%M:%S')
                supposed_time = ws2.cell(row=request_reader["row"],column=7).value
                delta = supposed_time - now
                if delta.days < 0:
                    print ("【过期还书！过期次数加一！】")
                    overdue_times = request_reader["过期次数"] + 1
                    ws2.cell(row=request_reader["row"],column=11).value = overdue_times
                    if overdue_times >= 2:
                        print ("【目前过期次数为" + str(overdue_times) + "次，借书权限暂停！】")
                        ws2.cell(row=request_reader["row"],column=10).value = "暂停"
                    else:
                        print ("【目前过期次数为" + str(overdue_times) + "次，请按时还书！】")                  
                ws2.cell(row=request_reader["row"],column=9).value = request_reader["借书记录"] + "," + isbn + "[" + now_str + "]"
                ws2.cell(row=request_reader["row"],column=5).value = None
                ws2.cell(row=request_reader["row"],column=6).value = None
                ws2.cell(row=request_reader["row"],column=7).value = None
                ws2.cell(row=request_reader["row"],column=8).value = now
                wr2.save(reader)                
                ws.cell(row=request_book["row"],column=14).value += "\n          " + "(" + now_str + ")" + request_reader["借书号"]  + request_reader["单位"] + request_reader["姓名"] + "[还书]"
                wr.save(library)               
                read_bookinfo()
                read_readerinfo()
                print ("\n【还书成功！】")
                input(header2+"\n点击回车键确认退出。")
                readerid = "0"
            else:
                readerid = "0"
        except Exception:
            print ("\n【读者借书号不存在，请重新输入！】")
            print (header2)
            readerid = input_request("请输入读者【借书号】进行查询，退出还书请按【0】\n读者借书号:")

def print_request_book(request, all_or_not):
    ###打印书籍信息###
    available_number = request["馆藏本数"]
    history = request["借阅记录"]
    if history == None:
        pass
    else:
        while history.find("[") != -1:
            status = history[history.find("[")+1:history.find("[")+3]
            if status == "借书":
                available_number -= 1
            elif status == "还书" or status == "丢失":
                available_number += 1
            history = history[history.find("[")+5:]
    print ("书籍名称:", request["书籍名称"])
    print ("作者:", request["作者"])
    print ("出版社:", request["出版社"])
    print ("价格:", request["价格"])
    print ("馆藏本数:", request["馆藏本数"])
    print ("可借本书:", available_number)
    print ("书籍位置:", "【"+str(request["书籍位置"])+"】")
    if all_or_not == False:
        if request["借阅次数"] == 0:
            print ("借阅记录:","该书借阅次数:", request["借阅次数"])
        else:
            borrow_log = request["借阅记录"]
            log_table = []
            while borrow_log.find("(") != -1:
                log_table.append(borrow_log[borrow_log.find("("):borrow_log.find("]")+1])
                borrow_log = borrow_log[borrow_log.find("]")+1:]
            print ("借阅记录:","该书借阅次数:", request["借阅次数"])
            if len(log_table) >= 5:
                print ("         【早期记录已省略，仅显示最近5条记录，若要查询完整记录请进入管理员模式。】")
                for log in log_table[-5:]:
                    print ("         ", log)
            else:
                for log in log_table:
                    print ("         ", log)
    else:
        print ("借阅记录:", request["借阅记录"])

def request_book():
    ###返回检索书籍信息###
    print (header1)
    isbn = input_request("请输入【ISBN】进行查询，退出请按【0】\nISBN:")
    while isbn != "0":
        try:
            request = data_book[isbn]
            print_request_book(request, False)
            isbn = input_request(header2+"\n请输入【ISBN】进行查询，退出请按【0】\nISBN:")
        except Exception:
            print ("【ISBN或借阅记录中的读者借书号不存在，请重新输入！】")
            isbn = input_request(header2+"\n请输入【ISBN】进行查询，退出请按【0】\nISBN:")

def print_request_reader(request, all_or_not):
    ###打印读者信息###
    print ("姓名:", request["姓名"])
    print ("性别:", request["性别"])
    print ("单位:", request["单位"])
    if request["所借书目"] == None:
        print ("所借书目:", "暂无所借书目")
    else:
        print ("所借书目:", request["所借书目"], data_book[request["所借书目"]]["书籍名称"])
    if request["借书日期"] == None:
        print ("借书日期:")
        print ("应还日期:")
    else:
        print ("借书日期:", request["借书日期"].strftime('%Y-%m-%d %H:%M:%S'))
        print ("应还日期:", request["应还日期"].strftime('%Y-%m-%d %H:%M:%S'))
    if request["还书日期"] == None:
        print ("还书日期:")
    else:
        print ("还书日期:", request["还书日期"].strftime('%Y-%m-%d %H:%M:%S'))
    
    if request["借阅次数"] == 0:
        print ("借书记录:", "无借书记录")
    else:
        borrow_log = request["借书记录"].split(",")
        print ("借书记录:", "曾借阅图书" + str(request["借阅次数"]) + "本：")
        if all_or_not == False:
            log_table = []
            for book in borrow_log:
                if book.find("(") != -1:
                    log_table.append("         "+book[book.find("(")+1:-1]+" 借书 "+data_book[book[:book.find("(")]]["书籍名称"]+" 【"+str(data_book[book[:book.find("(")]]["书籍位置"])+"】")
                elif book.find("{") != -1:
                    log_table.append("         "+book[book.find("{")+1:-1]+" 丢失 "+data_book[book[:book.find("{")]]["书籍名称"]+" 【"+str(data_book[book[:book.find("{")]]["书籍位置"])+"】")
                elif book.find("[") != -1:
                    log_table.append("         "+book[book.find("[")+1:-1]+" 还书 "+data_book[book[:book.find("[")]]["书籍名称"]+" 【"+str(data_book[book[:book.find("[")]]["书籍位置"])+"】")
            if len(log_table) >= 5:
                print ("         【早期记录已省略，仅显示最近5条记录，若要查询完整记录请进入管理员模式。】")
                for log in log_table[-5:]:
                    print (log)
            else:
                for log in log_table:
                    print (log)
        else:
            for book in borrow_log:
                if book.find("(") != -1:
                    print ("         "+book[book.find("(")+1:-1], "借书", data_book[book[:book.find("(")]]["书籍名称"], "【"+str(data_book[book[:book.find("(")]]["书籍位置"])+"】")
                elif book.find("{") != -1:
                    print ("         "+book[book.find("{")+1:-1], "丢失", data_book[book[:book.find("{")]]["书籍名称"], "【"+str(data_book[book[:book.find("{")]]["书籍位置"])+"】")
                else:
                    print ("         "+book[book.find("[")+1:-1], "还书", data_book[book[:book.find("[")]]["书籍名称"], "【"+str(data_book[book[:book.find("[")]]["书籍位置"])+"】")        
    print ("借书权限:", request["借书权限"])
    print ("过期次数:", request["过期次数"])

def request_reader():
    ###返回检索读者信息###
    print (header1)
    readerid = input_request("请输入读者【借书号】进行查询，退出请按【0】\n读者借书号:")
    while readerid != "0":
        try:
            request = data_reader[readerid]
            print_request_reader(request, False)
            readerid = input_request(header2+"\n请输入读者【借书号】进行查询，退出请按【0】\n读者借书号:")
        except Exception:
            print ("【读者借书号或借阅记录中的书目不存在，请重新输入！】")
            readerid = input_request(header2+"\n请输入读者【借书号】进行查询，退出请按【0】\n读者借书号:")

def reader_access_revise():
    ###修改读者借书权限###
    print (header1)
    readerid = input_request("请输入读者【借书号】，退出请按【0】\n读者借书号:")
    while readerid != "0":
        try:
            request_reader = data_reader[readerid]
            print_request_reader(request_reader, False)
            request = input_request(header2+"\n开通读者借书权限请按【1】\n暂停读者借书权限请按【2】\n重置读者过期次数请按【3】\n退出请按【0】\n")
            while request != "0":
                if request == "1":
                    ws2.cell(row=request_reader["row"],column=10).value = "开通"
                    wr2.save(reader)
                    read_readerinfo()
                    print ("【" + request_reader["单位"]+request_reader["姓名"]+"借书权限开通成功！】")
                    request = input_request(header2+"\n开通读者借书权限请按【1】\n暂停读者借书权限请按【2】\n重置读者过期次数请按【3】\n退出请按【0】\n")
                elif request == "2":
                    ws2.cell(row=request_reader["row"],column=10).value = "暂停"
                    wr2.save(reader)
                    read_readerinfo()
                    print ("【" + request_reader["单位"]+request_reader["姓名"]+"借书权限暂停成功！】")
                    request = input_request(header2+"\n开通读者借书权限请按【1】\n暂停读者借书权限请按【2】\n重置读者过期次数请按【3】\n退出请按【0】\n")
                elif request == "3":
                    ws2.cell(row=request_reader["row"],column=11).value = None
                    wr2.save(reader)
                    read_readerinfo()
                    print ("【" + request_reader["单位"]+request_reader["姓名"]+"过期次数重置成功！】")
                    request = input_request(header2+"\n开通读者借书权限请按【1】\n暂停读者借书权限请按【2】\n重置读者过期次数请按【3】\n退出请按【0】\n")
                else:
                    print ("\n【错误代码，请重新输！】")
                    request = input_request(header2+"\n开通读者借书权限请按【1】\n暂停读者借书权限请按【2】\n重置读者过期次数请按【3】\n退出请按【0】\n")
            return
        except Exception:
            print ("\n【读者借书号不存在，请重新输入！】\n"+header2)
            readerid = input_request("请输入读者【借书号】，退出还按【0】\n读者借书号:")

def book_info_entry_single():
    ###单本录入图书信息###
    global wr,ws
    print(header1+"\n请输入书籍ISBN、书籍本数、书籍位置，退出请输入【0】：")
    isbn = input("ISBN：")
    if isbn == "0":
        return
    number = input("本数：")
    if number == "0":
        return
    if not number.isdigit():
        number = "00"
    position = input("位置：")
    if position == "0":
        return
    while isbn != "0" and number != "0":
        print ("----------------------------")
        print ("【正在自动生成书籍信息，请稍后！】\n"+header2)
        bookinfo = writer(isbn, number, position)
        if type(bookinfo) == str:
            print (bookinfo)
            book_info_entry_single()
            return
        if str(bookinfo[0]) in data_book.keys() :
            print ("【书目已存在于数据库中】，信息如下：")
            print ("书籍名称:", data_book[str(bookinfo[0])]["书籍名称"])
            print ("馆藏本数:", data_book[str(bookinfo[0])]["馆藏本数"])
            print ("书籍位置:", data_book[str(bookinfo[0])]["书籍位置"])
            print ("============================")
            confirm = input_request("【是否与现存书目信息合并】？\n现存书目将增加【"+str(bookinfo[8])+"】本，书籍位置依然为【"+str(data_book[str(bookinfo[0])]["书籍位置"])+"】？\n确认请按【1】，取消请按【0】")
            while confirm != "1" and confirm != "0":
                confirm = input_request("确认请按【1】，取消请按【0】")
            if confirm == "1":
                excel_to_json()
                ws.cell(row=data_book[str(bookinfo[0])]["row"],column=9).value += bookinfo[8]
                wr.save(library)
                wr = load_workbook(library)
                ws = wr.worksheets[0]
                read_bookinfo()
                print ("【信息录入完成！】")
        else:
            excel_to_json()
            existed = ws.max_row
            for i in range(15):
                ws.cell(row=existed+1,column=i+1).value = bookinfo[i]
            wr.save(library)      
            wr = load_workbook(library)
            ws = wr.worksheets[0]
            read_bookinfo()
            print ("【信息录入完成！】")
        print(header2+"\n请输入书籍ISBN、书籍本数、书籍位置，退出请输入【0】：")
        isbn = input("ISBN：")
        if isbn == "0":
            return
        number = input("本数：")
        if number == "0":
            return
        if not number.isdigit():
            number = "00"
        position = input("位置：")
        if position == "0":
            return

def reader_reset_batch():
    ###重置读者信息###
    nrows = ws2.max_row
    for row in range(2,nrows+1):
        ws2.cell(row=row,column=5).value = None
        ws2.cell(row=row,column=6).value = None
        ws2.cell(row=row,column=7).value = None
        ws2.cell(row=row,column=8).value = None
        ws2.cell(row=row,column=9).value = None
        ws2.cell(row=row,column=10).value = "开通"
        ws2.cell(row=row,column=11).value = None
        ws2.cell(row=row,column=12).value = None
    wr2.save(reader)

def books_reset_batch():
    ###重置图书信息###
    nrows = ws.max_row
    for row in range(2,nrows+1):
        ws.cell(row=row,column=13).value = None
        ws.cell(row=row,column=14).value = None
    wr.save(library)  

def admin():
    ###管理员模块###
    global pw_admin, wr2, ws2
    instruction = "\n【管理员】请按指示进行相关操作：\n修改读者权限请按【1】\n单本录入图书请按【2】\n图书信息重置请按【3】\n读者信息重置请按【4】\n自动生成借书号请按【5】\n查看统计信息请按【6】\n设置还书期限请按【7】\n重置软件请按【8】\n查询书目完整信息请按【9】\n查询读者完整信息请按【10】\n恢复文件请按【11】\n退出请按【0】\n"
    print (header1)
    password = input("请输入【管理员密码】！退出请按【0】\n密码:") 
    if password == "0":
        return
    while password != pw_admin:
        password = input("【密码错误！】\n请输入密码！退出请按【0】\n密码:")
        if password == "0":
            return
    content = input_request(header1+instruction)
    while content != "0":
        if content == "1":
            excel_to_json()
            reader_access_revise()
            content = input_request(header1+instruction)
        elif content == "2":
            excel_to_json()
            book_info_entry_single()
            content = input_request(header1+instruction)
        elif content == "3":
            print (header1)
            confirm = input_request("该操作将清空所有图书借还书记录，确认请按【1】，取消请按【0】")
            while confirm != "1" and confirm != "0":
                confirm = input_request("该操作将清空所有图书借还书记录，确认请按【1】，取消请按【0】")
            if confirm == "1":
                excel_to_json()
                books_reset_batch()
                read_bookinfo()
                print ("【操作完成，图书信息重置成功！】")
            content = input_request(header1+instruction)
        elif content == "4":
            print (header1)
            confirm = input_request("该操作将清空所有借还书记录，并开通所有读者借书权限，确认请按【1】，取消请按【0】")
            while confirm != "1" and confirm != "0":
                confirm = input_request("该操作将清空所有借还书记录，并开通所有读者借书权限，确认请按【1】，取消请按【0】")
            if confirm == "1":
                excel_to_json()
                reader_reset_batch()
                read_readerinfo()
                print ("【操作完成，读者信息重置成功！】")
            content = input_request(header1+instruction)
        elif content == "5":
            excel_to_json()
            reader_id_generater()
            wr2 = load_workbook(reader)
            ws2 = wr2.worksheets[0]
            read_readerinfo()
            print ("【操作完成，读者借书号成功生成！】")
            content = input_request(header1+instruction)
        elif content == "6":
            excel_to_json()
            info_summary(data_book)
            content = input_request(header1+instruction)
        elif content == "7":
            print (header1)
            student_days = input("【学生】借书期限：")
            teacher_days = input("【教师】借书期限：")
            confirm = input_request("该操作将设置读者借书期限，但【不会】影响图书和读者信息。\n确认请按【1】，取消请按【0】")
            while confirm != "1" and confirm != "0":
                confirm = input_request("该操作将设置读者借书期限，但【不会】影响图书和读者信息。\n确认请按【1】，取消请按【0】")
            if confirm == "1":
                with open('meta_data.json', "r") as f1:
                    json_data = json.load(f1)
                json_data["student_days"] = student_days
                json_data["teacher_days"] = teacher_days
                with open('meta_data.json', "w") as f2:
                    json.dump(json_data,f2)
                print ("读者借书期限设置成功！")
            content = input_request(header1+instruction)
        elif content == "8":
            print (header1)
            confirm = input_request("该操作将重置软件登录信息及密码，但【不会】影响图书和读者信息。\n确认请按【1】，取消请按【0】")
            while confirm != "1" and confirm != "0":
                confirm = input_request("该操作将重置软件登录信息及密码，但【不会】影响图书和读者信息。\n确认请按【1】，取消请按【0】")
            if confirm == "1":
                print (header2)
                with open('meta_data.json', "r") as f1:
                    json_data = json.load(f1)
                json_data["status"] = "0"
                with open('meta_data.json', "w") as f2:
                    json.dump(json_data,f2)
                welcome, pw, pw_admin = initiallize()
                return ("initiallized")
            content = input_request(header1+instruction)
        elif content == "9":
            print (header1)
            isbn = input_request("请输入【ISBN】进行查询，退出请按【0】\nISBN:")
            while isbn != "0":
                try:
                    request = data_book[isbn]
                    print_request_book(request, True)
                    isbn = input_request(header2+"\n请输入【ISBN】进行查询，退出请按【0】\nISBN:")
                except Exception:
                    print ("【ISBN或借阅记录中的读者借书号不存在，请重新输入！】")
                    isbn = input_request(header2+"\n请输入【ISBN】进行查询，退出请按【0】\nISBN:")
            content = input_request(header1+instruction)
        elif content == "10":
            print (header1)
            readerid = input_request("请输入读者【借书号】进行查询，退出请按【0】\n读者借书号:")
            while readerid != "0":
                try:
                    request = data_reader[readerid]
                    print_request_reader(request, True)
                    readerid = input_request(header2+"\n请输入读者【借书号】进行查询，退出请按【0】\n读者借书号:")
                except Exception:
                    print ("【读者借书号或借阅记录中的书目不存在，请重新输入！】")
                    readerid = input_request(header2+"\n请输入读者【借书号】进行查询，退出请按【0】\n读者借书号:")
            content = input_request(header1+instruction)
        elif content == "11":
            print (header1)
            print ("正在恢复文件中，请稍后...")
            json_to_excel()
            print ("文件恢复完成，请返回根目录，并按文件名中的提示操作。")
            input ("点击回车键确认退出。")
            content = input_request(header1+instruction)
        else:
            print("\n【错误代码，请重新输！】")
            content = input_request(header1+instruction)

def initiallize():
    ###初始化###
    with open('meta_data.json', "r") as f1:
        json_data = json.load(f1)
    if json_data["status"] == "0":
        print ("【软件初始化】，请按提示输入相应内容！")
        json_data["status"] = "1"
        json_data["institution"] = input("【学校/机构名称】：")
        json_data["password"] = input("【登陆密码】：")
        json_data["administrator"] = input("【管理员密码】：")
        with open('meta_data.json', "w") as f2:
            json.dump(json_data,f2)
    return (json_data["institution"], json_data["password"], json_data["administrator"])

def excel_to_json():
    ###数据备份模块###
    def excel_to_json_reader(file_name):
        convert_dict = copy.deepcopy(data_reader)
        for reader in convert_dict:
            for key in ["借书日期", "应还日期", "还书日期"]:
                if convert_dict[reader][key] != None:
                    convert_dict[reader][key] = convert_dict[reader][key].strftime('%Y-%m-%d %H:%M:%S')
        with open(file_name + '_backup.json',"w") as f:
            json.dump(convert_dict,f)
    def excel_to_json_library(file_name):
        with open(file_name + '_backup.json',"w") as f:
            json.dump(data_book,f)
    excel_to_json_reader(reader)
    excel_to_json_library(library)

def json_to_excel():
    ###数据还原模块### 
    def json_to_excel_library(file_name):
        wb = Workbook()
        ws0 = wb.worksheets[0]
        ws0.title = u"书籍信息"
        ws1 = wb.create_sheet()
        ws1.title = u'书籍丢失信息'
        title = ["ISBN", "书籍名称", "作者", "出版社", "出版日期", "页数", "价格", "主题", "馆藏本数", "索书号", "内容简介", "信息来源", "借阅次数", "借阅记录", "书籍位置"]
        title2 = ["单位", "姓名", "性别", "ISBN", "书名", "价格", "登记时间"]
        for i in range(len(title)):
            ws0.cell(row=1,column=i+1).value = title[i]
        for i in range(len(title2)):
            ws1.cell(row=1,column=i+1).value = title2[i]
        with open(file_name, "r") as f:
            json_data = json.load(f)
            for book in json_data:
                book_single = [json_data[book][key] for key in title]
                existed = ws0.max_row
                for i in range(len(title)):
                    ws0.cell(row=existed+1,column=i+1).value = book_single[i]
        wb.save(file_name[:file_name.find(".")]+"（恢复文件，请将括号连带此提示删除并替换损坏文件即可使用）.xlsx")  
    
    def json_to_excel_reader(file_name):
        wb = Workbook()
        ws0 = wb.worksheets[0]
        ws0.title = u"读者信息"
        ws1 = wb.create_sheet()
        ws1.title = u'借阅记录'     
        title = ["借书号", "姓名", "性别", "单位", "所借书目", "借书日期", "应还日期", "还书日期", "借书记录", "借书权限", "过期次数", "借阅次数"]
        title2 = ["时间", "单位", "姓名", "借书号", "动作", "ISBN", "书名", "书籍位置"]
        for i in range(len(title)):
            ws0.cell(row=1,column=i+1).value = title[i]
        for i in range(len(title2)):
            ws1.cell(row=1,column=i+1).value = title2[i]
        with open(file_name, "r") as f:
            json_data = json.load(f)
            for reader in json_data:
                reader_single = []
                for key in title:
                    if key in ["借书日期", "应还日期", "还书日期"]:
                        if json_data[reader][key] != None:
                            d = datetime.datetime.strptime(json_data[reader][key], '%Y-%m-%d %H:%M:%S')
                        else:
                            d = None
                        reader_single.append(d)
                    else:
                        reader_single.append(json_data[reader][key])
                existed = ws0.max_row
                for i in range(len(title)):
                    ws0.cell(row=existed+1,column=i+1).value = reader_single[i]
        wb.save(file_name[:file_name.find(".")]+"（恢复文件，请将括号连带此提示删除并替换损坏文件即可使用）.xlsx")  
    json_to_excel_library('图书馆信息.xlsx_backup.json')
    json_to_excel_reader('读者信息.xlsx_backup.json')

def main():
    ###主界面###
    instruction = "\n请按指示进行相关操作：\n借书请按【1】\n还书请按【2】\n查询书目信息请按【3】\n查询读者信息请按【4】\n管理各类信息请按【5】\n帮助请按【6】\n退出请按【0】\n"
    print ("请确认“图书馆信息.xlsx”和“读者信息.xlsx”文件保持关闭状态，并与该软件置于同一目录下！")
    excel_to_json()
    print (header1)
    print ("欢迎进入"+ welcome + "图书馆管理系统！")
    print (header1)
    password = input("请输入密码！退出请按【0】\n密码:")
    if password == "0":
        return
    while password != pw:
        password = input("\n【密码错误！】\n请输入密码！退出请按0\n密码:")
        if password == "0":
            return
    print (header1)    
    print ("【密码正确，欢迎使用！】")
    print (header2)
    summary()
    print ("图书馆现存图书【"+str(summary2)+"】种，共计图书【"+str(summary1)+"】册，注册读者【"+str(summary3)+"】人。")
    print ("学生借书期限【"+ supposed_return_days_students + "】天，教师借书期限【" + supposed_return_days_teachers + "】天。")
    content = input_request(header1+instruction)
    while content != "0":
        if content == "1":
            excel_to_json()
            borrow_book()
            content = input_request(header1+instruction)
        elif content == "2":
            excel_to_json()
            return_book()
            content = input_request(header1+instruction)
        elif content == "3":
            request_book()
            content = input_request(header1+instruction)
        elif content == "4":
            request_reader()
            content = input_request(header1+instruction)
        elif content == "5":
            if admin() == "initiallized":
                input("软件重置成功，请点击回车键后重新打开软件！")
                return
            content = input_request(header1+instruction)
        elif content == "6":
            print (header2 + "\n感谢使用此套图书管理系统！\n软件作者：宋嘉勋\n联系方式：E-mail:jiaxun.song@outlook.com | Phone:15201410329 | Wechat:360911702")
            content = input_request(header1+instruction)
        else:
            print ("\n【错误代码，请重新输！】")
            content = input_request(header1+instruction)

welcome, pw, pw_admin = initiallize()
read_bookinfo()
read_readerinfo()
main()
