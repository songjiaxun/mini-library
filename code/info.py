# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import json
import globalvar
import copy
from validation import Validation

validation = Validation()

class Info(object):
    """Info read/write class"""    
    def __init__(self,libFile="../data/图书馆信息.xlsx",readerFile="../data/读者信息.xlsx"):
        ###
        ##数据初始化
        ###
        self.libFile = libFile
        self.readerFile = readerFile
        self.dataBook = {}
        self.dataReader = {}
        self.bookTotal= 0 #书籍总量（本）
        self.bookKinds= 0 #书籍种类总量（种）
        self.readerTotal =0 #读者数量 （人）
        ###从两个excel文档和一个json文档中载入初始数据###
        try:
            libWorkbook = load_workbook(libFile)
            self.books = libWorkbook.worksheets[0]#读取书籍信息
            self.booksLost = libWorkbook.worksheets[1]#读取书籍遗失

        except Exception:
            input("请确认“图书馆信息.xlsx”文件保持关闭状态，并置于 data 文件夹下！")                        
        try:
            readerWorkbook = load_workbook(readerFile)
            self.readers = readerWorkbook.worksheets[0]#读取读者信息
            self.logBorrow = readerWorkbook.worksheets[1]#读取借阅记录
        except Exception:
            input("请确认“读者信息.xlsx”文件保持关闭状态，并置于 data 文件夹下！")
        ###从 meta_data.json 中读取配置信息###
        try:
            with open('../data/meta_data.json',"r") as file:
                meta_data = json.load(file)
                self.supposedReturnDays_students = meta_data["student_days"]
                self.supposedReturnDays_teachers = meta_data["teacher_days"]
        except Exception:
            input("请确认“meta_data.json”文件保持关闭状态，并置于 data 文件夹下！")

        ##统计书籍数据##    
        self.bookRead() 
        #统计读者数据##
        self.readerRead()
    def bookRead(self):
        ###从excel文件读取图书信息,即原版的read_bookinfo()###
        books = self.books
        rows = books.max_row
        for book in range(2, rows+1):
            if books.cell(row=book,column=13).value == None:
                borrowedTimes = 0
            else:
                borrowedTimes = books.cell(row=book,column=13).value
            if not str(books.cell(row=book,column=1).value).isdigit():
                isbn = "0"
            else:
                isbn = str(books.cell(row=book,column=1).value)
            if not str(books.cell(row=book,column=9).value).isdigit():
                amount = 0
            else:
                amount = int(books.cell(row=book,column=9).value)
            infoBook = {
                "row": book,
                "ISBN": isbn,
                "书籍名称": books.cell(row=book,column=2).value,
                "作者": books.cell(row=book,column=3).value,
                "出版社": books.cell(row=book,column=4).value,
                "出版日期": books.cell(row=book,column=5).value,
                "页数": books.cell(row=book,column=6).value,
                "价格": books.cell(row=book,column=7).value,
                "主题": books.cell(row=book,column=8).value,
                "馆藏本数": amount,
                "索书号": books.cell(row=book,column=10).value,
                "书籍位置": books.cell(row=book,column=15).value,
                "借阅次数": borrowedTimes,
                "借阅记录": books.cell(row=book,column=14).value,
                "内容简介": books.cell(row=book,column=11).value,
                "信息来源": books.cell(row=book,column=12).value
            }                         
            self.dataBook[isbn] = infoBook
    def readerRead(self):
        ###从excel文件读取读者信息,即原版的reader_readerinfo()###
        readers = self.readers 
        rows = readers.max_row
        for reader in range(2, rows+1):
            if readers.cell(row=reader,column=11).value == None:
                count = 0
            else:
                count = readers.cell(row=reader,column=11).value
            if readers.cell(row=reader,column=12).value == None:
                borrowedTimes = 0
            else:
                borrowedTimes = readers.cell(row=reader,column=12).value
            if not str(readers.cell(row=reader, column=1).value).isdigit():
                readerId= 0
            else:
                readerId = str(readers.cell(row=reader,column=1).value)
            infoReader = {
                "row": reader,
                "借书号": readerId,
                "姓名": readers.cell(row=reader,column=2).value,
                "性别": readers.cell(row=reader,column=3).value,
                "单位": readers.cell(row=reader,column=4).value,
                "所借书目": readers.cell(row=reader,column=5).value,
                "借书日期": readers.cell(row=reader,column=6).value,
                "应还日期": readers.cell(row=reader,column=7).value,
                "还书日期": readers.cell(row=reader,column=8).value,
                "借书记录": readers.cell(row=reader,column=9).value,
                "借书权限": readers.cell(row=reader,column=10).value,
                "过期次数": count,
                "借阅次数": borrowedTimes
            }                         
            self.dataReader[readerId] = infoReader
    def reader_Write2Json(self,file):
        #备份读者信息至 json 文件，即原版 excel_to_json() 的拆分
        data_reader_bak = copy.deepcopy(self.data_reader)
        for key in data_reader_bak:
            for col in ['借书日期','应还日期','还书日期']:
                if data_reader_bak[key][col] != None:
                    data_reader_bak[key][col] = data_reader_bak[key][col].strftime('%Y-%m-%d %H:%M:%S')
        with open(file + '_bak.json','w') as f:
            json.dump(data_reader_bak,f)
    def book_Write2Json(self,file):
        #备份书籍信息至 json 文件，即原版 excel_to_json() 的拆分
        with open(file + '_bak.json','w') as f:
            json.dump(self.data_book,f)
    def bookBorrow(self):
        #书籍借阅操作#
        readerId = validation.inputs('请输入读者【借书号】，退出请按【0】\n读者借书号：')

        while readerId != '0':
            print(globalvar.border1)            
            try:
                req = self.dataReader[readerId]
                self.readerPrint(req,False)
                if req['借书权限'] != '开通':
                    print(globalvar.border2 + '\n【该读者暂无借书权限！】')
                    input(globalvar.border2 + '\n点击回车键确认退出借书。')
                    return 
                if req['所借书目'] != None:
                    print(globalvar.border2 + '\n【请将已借数目归还后再借新书！】')
                    input(globalvar.border2 + '\n点击回车键确认退出借书。')
                    return 
                print(globalvar.border2)
                isbn = validation.inputs('请输入欲借书目【ISBN】，退出借书请按【0】\nISBN：')
                if isbn == '0':
                    return 
                while isbn != '0':
                    try:
                        req_book = self.dataBook[isbn]
                        self.bookPrint(req_book, False)
                    except Exception:
                        print('\n【ISBN不存在，请重新输入！】')
                        print(globalvar.border2)
                        isbn = validation.inputs('请输入欲借书目【ISBN】，退出借书请按【0】\nISBN：')
                        if isbn == '0':
                            return
            except Exception:
                print('\n【读者借书号不存在，请重新输入！】')
                print(globalvar.border2)
                readerId = validation.inputs('请输入读者【借书号】，退出请按【0】\n 读者借书号：')
    def readerPrint(self, req, entries=False):
        ##打印读者信息##
        ##req:检索到的读者信息集##
        ##entries:是否显示全部借阅记录##
        print('姓名：', req['姓名'])
        print('性别：', req['性别'])
        print('单位：', req['单位'])
        if req['所借书目'] == None:
            print('所借书目：','暂无所借书目')
        else:
            print('所借书目：', req['所借书目'])
        if req['借书日期'] == None:
            print('借书日期：')
            print('应还日期：')
        else:
            print('借书日期：', req['借书日期'].strftime('%Y-%m-%d %H:%M:%S'))
            print('应还日期：', req['应还日期'].strftime('%Y-%m-%d %H:%M:%S'))
        
        if req['还书日期'] == None:
            print('还书日期：')
        else:
            print('还书日期：', req['还书日期'].strftime('%Y-%m-%d %H:%M:%S'))

        if req['借阅次数'] == 0:
            print('借书记录：','无借书记录')
        else:
            print('借书记录：','开发中...')
            # borrow_log = req_reader['借书记录'].split(',')
            # print('借书记录：', '曾借阅图书' + str(req_reader['借阅次数']) + '本：')
            # if entries == False:
            #     entry_list = []##存储借书记录
            #     for log in borrow_log:
            #         if log.find('(') != -1:
            #             entry_list.append('         '+log[log.find('(')+1:-1] + ' 借书 '+ log[:log.find('(')]['书籍名称'])
                        
            #     if len(entry_list) >= 5:
            #         print('【早期记录已省略，仅显示最近5条记录，若要查询完整记录请进入管理员模式。】')
            #         for entry in entry_list[-5:]:
            #             print(entry)
            #     else:
            #         for entry in entry_list:
            #             print(entry)
        print('借书权限：',req['借书权限'])
        print('过期次数:', req['过期次数'])

    def bookPrint(self, req_book, all_or_not):
        ###打印书籍信息###

        bookCollection = req_book['馆藏本数']
        bookLog = req_book['借阅记录']

        if bookLog == None:
            pass
        else:
            print(len(bookLog))
            # while bookLog.find('[') != -1:
            #     status = bookLog[bookLog.find('[')+1:bookLog.find(']')]
            #     print(status)
            #     if status == '借书':
            #         bookCollection -= 1
            #     elif status == '还书' or status == '丢失':
            #         bookCollection += 1
            #         print(222)
        print(111)

    def summary(self):
        ##统计馆藏数量、注册读者数等信息##

        for book in self.dataBook:
            if self.dataBook[book]['馆藏本数'] == None:
                pass
            else:
                self.bookTotal += self.dataBook[book]['馆藏本数']
                self.bookKinds += 1
        for reader in self.dataReader:
            self.readerTotal += 1




